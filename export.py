import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os


COLUMNS = [
    "company",
    "role",
    "location",
    "source",
    "apply_link",
    "salary",
    "visa_sponsorship",
    "posted_date",
    "match_score",
    "status",
    "applied_date",
    "notes",
]

HEADERS = [
    "Company",
    "Role",
    "Location",
    "Source",
    "Apply Link",
    "Salary",
    "Visa Sponsorship",
    "Posted Date",
    "Match Score",
    "Status",
    "Applied Date",
    "Notes",
]


def export_to_excel(jobs, output_dir):
    os.makedirs(output_dir, exist_ok=True)
    date_str = datetime.now().strftime("%Y-%m-%d")
    filepath = os.path.join(output_dir, f"jobs_{date_str}.xlsx")

    for job in jobs:
        job.setdefault("status", "Not Applied")
        job.setdefault("applied_date", "")
        job.setdefault("notes", "")
        job.pop("description_snippet", None)
        job.pop("visa_sponsorship_flag", None)

    df = pd.DataFrame(jobs)

    for col in COLUMNS:
        if col not in df.columns:
            df[col] = ""

    df = df[COLUMNS]
    df = df.sort_values(by=["match_score", "visa_sponsorship"], ascending=[False, True])
    df = df.reset_index(drop=True)

    df.to_excel(filepath, index=False, sheet_name="Jobs", engine="openpyxl")

    wb = load_workbook(filepath)
    ws = wb["Jobs"]

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    link_font = Font(color="0563C1", underline="single")

    visa_col = COLUMNS.index("visa_sponsorship") + 1
    link_col = COLUMNS.index("apply_link") + 1
    score_col = COLUMNS.index("match_score") + 1

    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        visa_cell = ws.cell(row=row, column=visa_col)
        visa_val = str(visa_cell.value).strip().lower()
        if visa_val == "yes":
            visa_cell.fill = green_fill
        elif visa_val == "not mentioned":
            visa_cell.fill = yellow_fill
        elif visa_val == "no":
            visa_cell.fill = red_fill

        link_cell = ws.cell(row=row, column=link_col)
        if link_cell.value and str(link_cell.value).startswith("http"):
            url = str(link_cell.value)
            link_cell.hyperlink = url
            link_cell.font = link_font
            link_cell.value = "Apply Here"

        score_cell = ws.cell(row=row, column=score_col)
        try:
            score = int(score_cell.value)
            if score >= 8:
                score_cell.fill = green_fill
            elif score >= 4:
                score_cell.fill = yellow_fill
        except (ValueError, TypeError):
            pass

    col_widths = {
        "company": 25,
        "role": 40,
        "location": 22,
        "source": 15,
        "apply_link": 15,
        "salary": 18,
        "visa_sponsorship": 18,
        "posted_date": 14,
        "match_score": 12,
        "status": 14,
        "applied_date": 14,
        "notes": 25,
    }
    for idx, col_name in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(idx)].width = col_widths.get(col_name, 15)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    _add_stats_sheet(wb, jobs)

    wb.save(filepath)
    print(f"\n  Excel saved: {filepath}")
    return filepath


def _add_stats_sheet(wb, jobs):
    ws = wb.create_sheet("Stats")

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    ws.cell(row=1, column=1, value="Metric").fill = header_fill
    ws.cell(row=1, column=1).font = header_font
    ws.cell(row=1, column=2, value="Count").fill = header_fill
    ws.cell(row=1, column=2).font = header_font

    total = len(jobs)
    nl_count = sum(1 for j in jobs if "netherlands" in j.get("location", "").lower()
                   or any(c in j.get("location", "").lower()
                          for c in ["amsterdam", "rotterdam", "eindhoven", "utrecht", "the hague"]))
    de_count = sum(1 for j in jobs if "germany" in j.get("location", "").lower()
                   or any(c in j.get("location", "").lower()
                          for c in ["berlin", "munich", "hamburg", "frankfurt", "cologne"]))
    visa_yes = sum(1 for j in jobs if j.get("visa_sponsorship") == "Yes")
    visa_maybe = sum(1 for j in jobs if j.get("visa_sponsorship") == "Not Mentioned")
    visa_no = sum(1 for j in jobs if j.get("visa_sponsorship") == "No")

    source_counts = {}
    for j in jobs:
        src = j.get("source", "Unknown")
        source_counts[src] = source_counts.get(src, 0) + 1

    stats = [
        ("Total Jobs Found", total),
        ("Netherlands", nl_count),
        ("Germany", de_count),
        ("", ""),
        ("Visa Sponsorship: Yes", visa_yes),
        ("Visa Sponsorship: Not Mentioned", visa_maybe),
        ("Visa Sponsorship: No", visa_no),
        ("", ""),
    ]
    for src, count in source_counts.items():
        stats.append((f"Source: {src}", count))

    for i, (metric, count) in enumerate(stats, 2):
        ws.cell(row=i, column=1, value=metric)
        ws.cell(row=i, column=2, value=count)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 12
